#! /usr/bin/python3
# -*- coding: UTF-8 -*-

'''

  Weekly report Generator for casa GZ MME-AMF team
  Author: sunjianhua
  TimeStamp:2022/9/27
  e-mail: 
          sunjianhua@casachina.com.cn

  Description:  The present script is able to:
                    1: Get data from the weeklyReport.xlsx exported by dingding
                    2: Output to another xlsx file in the format required by CasA MME-AMF Team

'''

import sys, os, time
from datetime import datetime
from optparse import OptionParser
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
#from openpyxl.writer.excel import ExcelWriter
from enum import Enum


# *************** constant Start *****************************
Spacing_between_tables = 3

DD_START_ROW = 3
DD_NAME_COLUMN = 2
DD_THIS_WEEK_TYPE_COLUMN = 6
DD_THIS_WEEK_ISSUE_PREFIX_COLUMN = 7
DD_THIS_WEEK_ISSUE_ID_COLUMN = 8
DD_THIS_WEEK_TASK_DESCRIPTION_COLUMN = 9
DD_THIS_WEEK_STATUS_COLUMN = 10
DD_THIS_WEEK_PROGRESS_COLUMN = 11
DD_NEXT_WEEK_TYPE_COLUMN = 12
DD_NEXT_WEEK_ISSUE_PREFIX_COLUMN = 13
DD_NEXT_WEEK_ISSUE_ID_COLUMN = 14
DD_NEXT_WEEK_TASK_DESCRIPTION_COLUMN = 15

WR_START_ROW = 2
WR_TIME_COLUMN = 1
WR_SUMMARY_COLUMN = 2
WR_ESTIMATE_TIME_COLUMN = 3
WR_OWNER_COLUMN = 4
WR_STATUS_COLUMN = 5
WR_REAL_WORKING_HOURS_COLUMN = 6
WR_COMMENT_COLUMN = 7

WR_TIME_COLUMN_C = 'A'
WR_SUMMARY_COLUMN_C = 'B'
WR_ESTIMATE_TIME_COLUMN_C = 'C'
WR_OWNER_COLUMN_C = 'D'
WR_STATUS_COLUMN_C = 'E'
WR_REAL_WORKING_HOURS_COLUMN_C = 'F'
WR_COMMENT_COLUMN_C = 'G'
        
ROW_HEIGHT_UNIT = 17
# *************** constant End   *****************************

class weeklyRptGenerator(object):
    logLevelEnum = Enum('logLevelEnum', ('ALL', 'DEBUG', 'INFO', 'WARN', 'ERROR'))
    
    #set default log level
    logLevel = logLevelEnum.DEBUG
    
    header = ["Time", "Summary", "Estimate time(hours)", "Owner", "Status", "Real Working hours", "Comment"]
    
    teamMateConfigFilePath = "./TeamMateConfig.txt"

    def __init__(self):
        self.teamMateList = []
        
        #reprot column for one person start
        self.summary = ''
        self.estimateTime = 40
        self.owner = ''
        self.status = ''
        self.realWorkingHours = 40
        self.comment = ''
        #reprot column for one person end

    def updateLogLevel(self, options):
        if options.logLevel == "ALL":
            self.logLevel = self.logLevelEnum.ALL
        elif options.logLevel == "DEBUG":
            self.logLevel = self.logLevelEnum.DEBUG
        elif options.logLevel == "INFO":
            self.logLevel = self.logLevelEnum.INFO
        elif options.logLevel == "WARN":
            self.logLevel = self.logLevelEnum.WARN
        elif options.logLevel == "ERROR":
            self.logLevel = self.logLevelEnum.ERROR
        
    def pLog(self, logLevel, log):

        line_num = sys._getframe().f_back.f_lineno
        file_name = sys._getframe().f_back.f_code.co_filename
        StyleTime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        prefix = "[{}][{}][{}]".format(StyleTime, file_name, line_num)

        if logLevel.value >= self.logLevel.value:
            logPrefix = "ALL"
            if logLevel == self.logLevelEnum.DEBUG:
                logPrefix = "DEBUG"
            elif logLevel == self.logLevelEnum.INFO:
                logPrefix = "INFO"
            elif logLevel == self.logLevelEnum.WARN:
                logPrefix = "WARN"
            elif logLevel == self.logLevelEnum.ERROR:    
                logPrefix = "ERROR"
            
            #comment out for windows
            #Enable color font
            #if logLevel == self.logLevelEnum.INFO:
            #    print ("\033[1;32;40m%s[%s]:%s\033[0m" % (prefix, logPrefix, log))  #green
            #elif logLevel == self.logLevelEnum.WARN:
            #    print ("\033[1;33;40m%s[%s]:%s\033[0m" % (prefix, logPrefix, log))  #yellow
            #elif logLevel == self.logLevelEnum.ERROR:
            #    print ("\033[1;31;40m%s[%s]:%s\033[0m" % (prefix, logPrefix, log))  #red
            #else:
            #    print ("%s[%s]:%s" % (prefix, logPrefix, log)) 
            print ("%s[%s]:%s" % (prefix, logPrefix, log)) 

    def initArgsAnalysis(self):
        #comment out for windows 
        #usage = """
        #\033[1;32;40mCasa Systems\033[0m
        
        #\033[0;32;40mDescription:  The present script is able to:
        #              1: Get data from the weeklyReport.xlsx exported by dingding
        #              2: Output to another xlsx file in the format required by CasA MME-AMF Team.\033[0m

        #Command example: %prog [options] arg
        #""" 
        
        usage = """
        Casa Systems
        
        Description:  The present script is able to:
                      1: Get data from the weeklyReport.xlsx exported by dingding
                      2: Output to another xlsx file in the format required by CasA MME-AMF Team.

        Command example: %prog [options] arg
        """ 
        
        parser = OptionParser(usage, version="%prog v1.0.0")
        
        #Common args
        parser.add_option("--logLevel", action="store", default="INFO", dest="logLevel", help="value of log level:ALL/DEBUG/INFO/WARN/ERROR")
        parser.add_option("-i","--inPutFile", action="store", default="", dest="inPutFile", help="The source xlsx file that data should come from")
        parser.add_option("-o","--outPutFile", action="store", default="./WeeklyReport.xlsx", dest="outPutFile", help="The target xlsx file that data should output to")
        parser.add_option("-n","--sheetName", action="store", default="MME", dest="sheetName", help="The target sheet that data should output to, eg:MME/AMF/...")
            
        return parser

    def checkOps(self, options):
        #check whether the file name is valid.
        
        #check the inPut xlsx file 
        if options.inPutFile == "" or not os.path.isfile(options.inPutFile):
            self.pLog(self.logLevelEnum.INFO, "inPutFile %s doesn't exist!" % options.inPutFile)
            return False
        
        #No need to check the outPut xlsx file,we will do it in function: outPutFileInit

        return True
        
    def loadCfg(self, options):
        result = False
        
        commentStart = False
        
        targetTeamName = options.sheetName
        
        targetTeamFound = False

        self.pLog(self.logLevelEnum.INFO, "Start reading %s team info from %s..." % (targetTeamName, self.teamMateConfigFilePath))

        teamMateConfig = open(self.teamMateConfigFilePath, 'r', encoding='utf-8')

        for line in teamMateConfig:
            if line.strip() == '' : 
                #blank string, continue
                continue

            line = line.strip().split(' ')

            if line[0].encode('utf-8').decode('utf-8') == '###':   #it's a block comment start/end
                if commentStart:
                    commentStart = False #block comment end
                else:
                    commentStart = True #block comment start
                    
                continue
                    
            if commentStart:
                #we are in the block comment
                continue

            if len(line) == 3 and line[0].encode('utf-8').decode('utf-8') == '#' and line[2].encode('utf-8').decode('utf-8') == 'team':
                #a new team info, like: # MME team
                #notice, a new team info start is also an old team info end.
                if targetTeamFound:
                    #target team info ends
                    self.pLog(self.logLevelEnum.INFO, "Complete.")
                    result = True
                    return result
                
                #Check whether it's the target team we want
                if line[1].encode('utf-8').decode('utf-8') == targetTeamName:
                    #traget team info found
                    targetTeamFound = True
                    continue
                
            #check whether we are reading the target team info: 
            if targetTeamFound and len(line) == 2:    
                chineseName = line[0].encode('utf-8').decode('utf-8')
                tmp_list = [chineseName, line[1]]
                self.teamMateList.append(tmp_list)
                self.pLog(self.logLevelEnum.INFO, "%s : %s" % (tmp_list[0], tmp_list[1]))

        if targetTeamFound: 
            self.pLog(self.logLevelEnum.INFO, "Complete.")
            result = True
        else:
            self.pLog(self.logLevelEnum.ERROR, "Missing the %s team teamamte list within %s." % (targetTeamName, self.teamMateConfigFilePath))
        
        return result

    def createXlsxFile(self, options):
        #create xlsx file
        wb = Workbook()
        
        #a new sheet will be created automatically:  sheet
        #sheetNames = wb.sheetnames
        #print(sheetNames)
        
        #Because we create a new work book,and it have a sheet named 'sheet' automatically, 
        #so we need to rename it to our target sheet in this case.
        #We don't want another empty sheet in the workbook. 
        ws = wb.active
        ws.title = options.sheetName
        
        #store it to disk
        wb.save(r'' + options.outPutFile)    
        
        wb.close()

    def outPutFileInit(self, options):
        #File part Start
        #check whether the outPut xlsx file exist,if not, create a new one.
        if not os.path.isfile(options.outPutFile):
            self.pLog(self.logLevelEnum.INFO, "outPutFile %s doesn't exist，creating..." % options.outPutFile)
            self.createXlsxFile(options)
        #File part End
        
        #Sheet part Start
        #check whether the target sheet of outPut xlsx file exist,
        #         Yes: Remove it and recreate
        #         No:  Create one

        #load outPutFile
        wb = load_workbook(r''+ options.outPutFile)

        if options.sheetName in wb.sheetnames:
           #remove it
           #note that, there must be one sheet at least for one workbook,
           #to avoid the error because of no sheet for one workbook,
           #we create a tmp sheet,and then remove the old sheet,and then rename the tmp sheet to our target sheet.
           wb.create_sheet("avoidNoSheetError")
           wb.remove(wb[options.sheetName])
           wb.save(r'' + options.outPutFile) 
        
        #create target sheet
        wb.create_sheet(options.sheetName)
        if "avoidNoSheetError" in  wb.sheetnames:
            #remove this,just to avoid no sheet for workbook error.
            wb.remove(wb["avoidNoSheetError"])
            
        #inject the header
        ws = wb[options.sheetName]
        ws.append(self.header)

        wb.save(r'' + options.outPutFile)
        #Sheet part End
    
        #close outPutFile
        wb.close()
        
        self.pLog(generator.logLevelEnum.INFO, "%s init complete." % options.outPutFile)

    def handleOnWeek(self, options, wb_o, ws_o, ws_i, rowIndex):
        #Leave the first three row blank
        wr_rowCursor_this_week_part = rowIndex + Spacing_between_tables
        wr_rowCursor_next_week_part = wr_rowCursor_this_week_part + len(self.teamMateList) + Spacing_between_tables
        
        #cells merge for time
        ##This week part
        tmpCellsMergeEndRow = wr_rowCursor_this_week_part + len(self.teamMateList) - 1
        ws_o.merge_cells(start_row = wr_rowCursor_this_week_part, end_row = tmpCellsMergeEndRow, start_column = 1, end_column = 1)
        ##Next week part
        tmpCellsMergeEndRow = wr_rowCursor_next_week_part + len(self.teamMateList) - 1
        ws_o.merge_cells(start_row = wr_rowCursor_next_week_part, end_row = tmpCellsMergeEndRow, start_column = 1, end_column = 1)
        wb_o.save(r'' + options.outPutFile)
        
        if ws_i.max_row < DD_START_ROW:
            self.pLog(self.logLevelEnum.ERROR, "Invalid intput xlsx file:%s" % options.inPutFile)
            return False
            
        for teamMate in self.teamMateList:
            chineseName = teamMate[0]
            self.owner = teamMate[1]
            
            row_height_t = 1  #row height for this week
            row_height_n = 1  #row height for next week
            column_width = 10
            
            #search the report for target teammate
            targetTeamMateFound = False
            task_for_this_week = ''
            task_for_next_week = ''
            idx_this_week = 1
            idx_next_week = 1
            for i in range(DD_START_ROW, ws_i.max_row + 1):
                tmpName = ws_i.cell(row = i, column = DD_NAME_COLUMN).value
                if targetTeamMateFound and tmpName != '':
                    #target teammate info ends,another teammate info found, but we don't need it here,just continue.
                    break
                    
                if tmpName != '' and tmpName.encode('utf-8').decode('utf-8') == chineseName.encode('utf-8').decode('utf-8'):    
                    targetTeamMateFound = True
                    
                if targetTeamMateFound:
                    #get info from input file
                    
                    #get this week
                    if len(ws_i.cell(row = i, column = DD_THIS_WEEK_TYPE_COLUMN).value.strip()) > 0:
                        t_t = '%d.[%s] ' % (idx_this_week, ws_i.cell(row = i, column = DD_THIS_WEEK_TYPE_COLUMN).value)
                    
                        self.pLog(self.logLevelEnum.DEBUG, "SJH %d DD_THIS_WEEK_ISSUE_PREFIX_COLUMN %d" % (i, len(ws_i.cell(row = i, column = DD_THIS_WEEK_ISSUE_PREFIX_COLUMN).value.strip())))
                        if len(ws_i.cell(row = i, column = DD_THIS_WEEK_ISSUE_PREFIX_COLUMN).value.strip()) > 0:
                            self.pLog(self.logLevelEnum.DEBUG, "SJH in")
                            t_t = t_t + ws_i.cell(row = i, column = DD_THIS_WEEK_ISSUE_PREFIX_COLUMN).value + '-' + str(ws_i.cell(row = i, column = DD_THIS_WEEK_ISSUE_ID_COLUMN).value) + ':'
                    
                        t_t = t_t + ws_i.cell(row = i, column = DD_THIS_WEEK_TASK_DESCRIPTION_COLUMN).value + '\n'
                        row_height_t += 1
                    
                        t_t = t_t + '    - Status: ' + ws_i.cell(row = i, column = DD_THIS_WEEK_STATUS_COLUMN).value + '\n'
                        row_height_t += 1
                    
                        self.pLog(self.logLevelEnum.DEBUG, "SJH %d DD_THIS_WEEK_PROGRESS_COLUMN %d" % (i, len(ws_i.cell(row = i, column = DD_THIS_WEEK_PROGRESS_COLUMN).value.strip())))
                        if  len(ws_i.cell(row = i, column = DD_THIS_WEEK_PROGRESS_COLUMN).value.strip()) > 0:
                            self.pLog(self.logLevelEnum.DEBUG, "SJH in")
                            t_t = t_t + '    - Progress: ' + ws_i.cell(row = i, column = DD_THIS_WEEK_PROGRESS_COLUMN).value + '\n'
                            row_height_t += 1
                        
                        task_for_this_week = task_for_this_week + t_t + '\n'
                        row_height_t += 1
                    
                        idx_this_week = idx_this_week + 1
                        
                    #get next week
                    t_n = ''
 
                    self.pLog(self.logLevelEnum.DEBUG, "SJH %d DD_NEXT_WEEK_TYPE_COLUMN %d" % (i, len(ws_i.cell(row = i, column = DD_NEXT_WEEK_TYPE_COLUMN).value.strip())))
                    if len(ws_i.cell(row = i, column = DD_NEXT_WEEK_TYPE_COLUMN).value.strip()) > 0:
                        self.pLog(self.logLevelEnum.DEBUG, "SJH in")
                        t_n = '%d.[%s] ' % (idx_next_week, ws_i.cell(row = i, column = DD_NEXT_WEEK_TYPE_COLUMN).value)
                        
                        self.pLog(self.logLevelEnum.DEBUG, "SJH %d DD_NEXT_WEEK_ISSUE_PREFIX_COLUMN %d" % (i, len(ws_i.cell(row = i, column = DD_NEXT_WEEK_ISSUE_PREFIX_COLUMN).value.strip())))
                        if len(ws_i.cell(row = i, column = DD_NEXT_WEEK_ISSUE_PREFIX_COLUMN).value.strip()) > 0:
                            self.pLog(self.logLevelEnum.DEBUG, "SJH in")
                            t_n = t_n + ws_i.cell(row = i, column = DD_NEXT_WEEK_ISSUE_PREFIX_COLUMN).value + '-' + str(ws_i.cell(row = i, column = DD_NEXT_WEEK_ISSUE_ID_COLUMN).value) + ':'

                        t_n = t_n + ws_i.cell(row = i, column = DD_NEXT_WEEK_TASK_DESCRIPTION_COLUMN).value + '\n'
                        row_height_n += 1                        
                        
                        task_for_next_week = task_for_next_week + t_n + '\n'
                        row_height_n += 1
                        
                        idx_next_week = idx_next_week + 1
            
            self.comment = task_for_this_week
            self.summary = task_for_next_week
            
            #store the target teammate info to output file
            ##This week part
            ws_o.row_dimensions[wr_rowCursor_this_week_part].height = row_height_t * ROW_HEIGHT_UNIT
            ws_o.cell(row = wr_rowCursor_this_week_part, column = WR_ESTIMATE_TIME_COLUMN).value = self.estimateTime
            ws_o.cell(row = wr_rowCursor_this_week_part, column = WR_OWNER_COLUMN).value = self.owner
            ws_o.cell(row = wr_rowCursor_this_week_part, column = WR_REAL_WORKING_HOURS_COLUMN).value = self.realWorkingHours
            ws_o.cell(row = wr_rowCursor_this_week_part, column = WR_COMMENT_COLUMN).value = self.comment
            wr_rowCursor_this_week_part = wr_rowCursor_this_week_part + 1
            
            
            ##Next week part
            ws_o.row_dimensions[wr_rowCursor_next_week_part].height = row_height_n * ROW_HEIGHT_UNIT
            ws_o.cell(row = wr_rowCursor_next_week_part, column = WR_SUMMARY_COLUMN).value = self.summary
            ws_o.cell(row = wr_rowCursor_next_week_part, column = WR_ESTIMATE_TIME_COLUMN).value = self.estimateTime
            ws_o.cell(row = wr_rowCursor_next_week_part, column = WR_OWNER_COLUMN).value = self.owner
            wr_rowCursor_next_week_part = wr_rowCursor_next_week_part + 1
            
        wb_o.save(r'' + options.outPutFile)
        
        self.pLog(self.logLevelEnum.INFO, "Data convert complete.")
    
    def setColumnWidth(self, options, wb_o, ws_o):
        ws_o.column_dimensions[WR_TIME_COLUMN_C].width = 15
        ws_o.column_dimensions[WR_SUMMARY_COLUMN_C].width = 60
        ws_o.column_dimensions[WR_ESTIMATE_TIME_COLUMN_C].width = 23
        ws_o.column_dimensions[WR_OWNER_COLUMN_C].width = 16
        ws_o.column_dimensions[WR_STATUS_COLUMN_C].width = 16
        ws_o.column_dimensions[WR_REAL_WORKING_HOURS_COLUMN_C].width = 20
        ws_o.column_dimensions[WR_COMMENT_COLUMN_C].width = 70
        wb_o.save(r'' + options.outPutFile)
        
    def setFromCenter(self, options, wb_o, ws_o):
        nRows = ws_o.max_row  #get num of row 
        nCols = ws_o.max_column #get num of column
        
        for i in range(nRows):
            for j in range(nCols):
                #For summary and comment,just wrapText, no set from center
                if (j + 1 == WR_SUMMARY_COLUMN and i+1 != 1) or (j + 1 == WR_COMMENT_COLUMN and i+1 != 1):
                    ws_o.cell(row = i + 1, column = j + 1).alignment = Alignment(wrapText= True)
                    continue
                ws_o.cell(row = i + 1, column = j + 1).alignment = Alignment(horizontal = 'center', vertical = 'center', wrapText= True)

        wb_o.save(r'' + options.outPutFile)

    def startDataConvert(self, options):
        #init outPutFile
        self.outPutFileInit(options)
        
        #load outPutFile
        wb_o = load_workbook(r''+ options.outPutFile)
        
        #load target sheet
        ws_o = wb_o[options.sheetName]
        
        #load inPutFile
        wb_i = load_workbook(r''+ options.inPutFile)
        
        ws_i = wb_i.active
        
        #set width for columns
        self.setColumnWidth(options, wb_o, ws_o)
        
        self.handleOnWeek(options, wb_o, ws_o, ws_i, WR_START_ROW)
        
        #adjust to set data from center
        self.setFromCenter(options, wb_o, ws_o)
        
        #wb close
        wb_o.close()
        wb_i.close()

if __name__ == "__main__":
    generator = weeklyRptGenerator()
    
    generator.pLog(generator.logLevelEnum.INFO, "Weekly Report Generator init...")
    
    parser = generator.initArgsAnalysis()    

    #Args analysis
    (options, args) = parser.parse_args()
    print(options)

    #update loglevel
    generator.updateLogLevel(options)    

    #check args
    opsValid = generator.checkOps(options)
    if not opsValid:
        exit()
    else:
        generator.pLog(generator.logLevelEnum.DEBUG, "Options/Arguments valid.")
    
    #load cfg for target team
    if not generator.loadCfg(options):
        exit()
    
    #start data handle
    generator.startDataConvert(options)
    
    exit()